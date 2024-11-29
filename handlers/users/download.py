import os

from aiogram import types
from aiogram.dispatcher import FSMContext
from asyncpg.pgproto.pgproto import timedelta

from loader import dp, db

import tempfile
from openpyxl.styles import Alignment
import openpyxl


async def download_questionnaires():
    questionnaires = await db.select_all_questionnaires()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = "Foydalanuvchi"
    worksheet['C1'] = "MTM"
    worksheet['D1'] = 'Vaqt'

    for cell in ['A1', 'B1', 'C1', 'D1']:
        worksheet[cell].alignment = Alignment(horizontal='center')

    worksheet.cell(row=1, column=1, value='T/r')
    worksheet.cell(row=1, column=2, value='Foydalanuvchi')
    worksheet.cell(row=1, column=3, value="MTM")
    worksheet.cell(row=1, column=4, value='Vaqt')

    tr = 0
    for questionnaire in questionnaires:
        tr += 1
        user_id = questionnaire['user_id']
        user = await db.select_user(user_id=user_id)
        full_name = user['full_name']
        time = questionnaire['created_at'] + timedelta(hours=5)

        worksheet.cell(row=tr + 1, column=1, value=tr).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=2, value=full_name).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=3, value=questionnaire['kindergarten']).alignment = Alignment(
            horizontal='center')
        worksheet.cell(row=tr + 1, column=4, value=time.strftime("%d.%m.%Y  %H:%M")).alignment = Alignment(
            horizontal='center')

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Questionnaire.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(text="Download", state="*")
async def get_branch(message: types.Message, state: FSMContext):
    try:
        await state.finish()
    except:
        pass

    telegram_id = message.from_user.id
    users = await db.select_users(telegram_id=telegram_id)
    user = users[0]
    role = user['role']
    if role == 'user':
        return

    temp_dir = await download_questionnaires()

    with open(os.path.join(temp_dir, 'Questionnaire.xlsx'), 'rb') as file:
        await message.answer_document(document=file)

    os.remove(os.path.join(temp_dir, 'Questionnaire.xlsx'))
